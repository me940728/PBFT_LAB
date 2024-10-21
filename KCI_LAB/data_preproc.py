import openpyxl
import json
import os
import glob

# 엑셀 파일 경로 설정 (origin_data 폴더 내 모든 xlsx 파일)
# file_path = '/Users/admin/Downloads/ll/envs/pbft/source/PBFT_LAB/KCI_LAB/origin_data/YELLOWDUST-1_1.xlsx' # 특정 파일 테스트
file_dir = '/Users/admin/Downloads/ll/envs/pbft/source/PBFT_LAB/KCI_LAB/origin_data'
xlsx_files = glob.glob(os.path.join(file_dir, '*.xlsx'))

# 전역 변수 정의
data_json = []


# 엑셀 파일 순회
for index, file_path in enumerate(xlsx_files, start=1):
    # 현재 작업 중인 파일 출력
    print(f"Processing file {index}/{len(xlsx_files)}: {file_path}")
    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Korean sentence 임시 저장 변수
    kor_sentence = None
    gloss_id_list = []
    start_list = []

    # 엑셀 순회
    row = 1
    while row <= sheet.max_row:
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        row_data = [val for val in row_data if val is not None]

        if not row_data:
            row += 1
            continue

        first_cell = row_data[0]

        if isinstance(first_cell, str) and first_cell == 'Korean sentence : ':
            # 이전에 저장된 데이터로 JSON 생성
            if kor_sentence is not None:
                # start_list와 gloss_id_list 정렬
                if start_list and gloss_id_list:
                    sorted_pairs = sorted(zip(start_list, gloss_id_list), key=lambda x: x[0])
                    sorted_start, sorted_gloss_id = zip(*sorted_pairs)
                else:
                    sorted_start, sorted_gloss_id = [], []
                # ksl 값 생성 (gloss_id를 문자열로 변환)
                ksl_value = ' '.join(map(str, sorted_gloss_id))
                # JSON 데이터 생성
                data_json.append({
                    'kor': kor_sentence,
                    'ksl': ksl_value,
                    'time': list(sorted_start)
                })
            # 새 Korean sentence 저장
            kor_sentence = row_data[1] if len(row_data) > 1 else None
            # start_list 초기화
            start_list = []
            gloss_id_list = []
            row += 1

        elif isinstance(first_cell, str) and first_cell.startswith('sign_gestures_'):
            # gloss_id 값 추가
            for value in row_data[1:]:
                if value not in [None, '', 'gloss_id : ']:
                    gloss_id_list.append(value)
            # 다음 행의 start 값 추가
            row += 1
            next_row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            next_row_data = [val for val in next_row_data if val not in [None, '', 'start(s) : ']]
            start_list.extend(next_row_data)
            # 다음 행을 건너뛰기
            row += 1
        else:
            row += 1

    # 마지막 Korean sentence에 대한 JSON 생성
    if kor_sentence is not None:
        # start_list와 gloss_id_list 정렬
        if start_list and gloss_id_list:
            sorted_pairs = sorted(zip(start_list, gloss_id_list), key=lambda x: x[0])
            sorted_start, sorted_gloss_id = zip(*sorted_pairs)
        else:
            sorted_start, sorted_gloss_id = [], []
        # ksl 값 생성 (gloss_id를 문자열로 변환)
        ksl_value = ' '.join(map(str, sorted_gloss_id))
        # JSON 데이터 생성
        data_json.append({
            'kor': kor_sentence,
            'ksl': ksl_value,
            'time': list(sorted_start)
        })

# JSON 파일 저장
output_dir = '/Users/admin/Downloads/ll/envs/pbft/source/PBFT_LAB/KCI_LAB/json_data'
os.makedirs(output_dir, exist_ok=True)
output_file_path = os.path.join(output_dir, 'output_data.json')
with open(output_file_path, 'w', encoding='utf-8') as json_file:
    json.dump(data_json, json_file, ensure_ascii=False, indent=4)

# 결과 출력
#for data in data_json:
#    print(data)
print("done!")